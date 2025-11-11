import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
import re
import threading

class MonotaroExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title('モノタロウ注文情報Excel作成')
        self.root.geometry('900x700')
        
        self.mode_var = tk.StringVar(value='new')
        self.data_list = []
        
        self.create_widgets()
        
    def create_widgets(self):
        # === モード選択フレーム ===
        mode_frame = ttk.LabelFrame(self.root, text='1. モード選択', padding=10)
        mode_frame.grid(row=0, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Radiobutton(mode_frame, text='新規作成', variable=self.mode_var, 
                       value='new', command=self.toggle_mode).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(mode_frame, text='既存ファイルに追加', variable=self.mode_var, 
                       value='append', command=self.toggle_mode).pack(side=tk.LEFT, padx=10)
        
        # === ファイル設定フレーム ===
        file_frame = ttk.LabelFrame(self.root, text='2. ファイル設定', padding=10)
        file_frame.grid(row=1, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        # 新規作成用
        self.new_file_frame = ttk.Frame(file_frame)
        self.new_file_frame.pack(side=tk.LEFT, padx=10)
        
        ttk.Label(self.new_file_frame, text='ファイルパス:').pack()
        file_path_frame = ttk.Frame(self.new_file_frame)
        file_path_frame.pack(pady=5)
        self.entry_file_path = ttk.Entry(file_path_frame, width=40)
        self.entry_file_path.pack(side=tk.LEFT, padx=5)
        ttk.Button(file_path_frame, text='参照', 
                  command=lambda: self.browse_file(self.entry_file_path, save=True)).pack(side=tk.LEFT)
        
        ttk.Label(self.new_file_frame, text='シート名:').pack(pady=(10, 0))
        self.entry_sheet_name = ttk.Entry(self.new_file_frame, width=20)
        self.entry_sheet_name.pack(pady=5)
        self.entry_sheet_name.insert(0, '注文内容')
        
        # 既存ファイル用
        self.append_file_frame = ttk.Frame(file_frame)
        
        ttk.Label(self.append_file_frame, text='ファイルパス:').pack()
        append_path_frame = ttk.Frame(self.append_file_frame)
        append_path_frame.pack(pady=5)
        self.entry_existing_file = ttk.Entry(append_path_frame, width=40, state='disabled')
        self.entry_existing_file.pack(side=tk.LEFT, padx=5)
        self.btn_browse_existing = ttk.Button(append_path_frame, text='参照', 
                                             command=lambda: self.browse_file(self.entry_existing_file, save=False),
                                             state='disabled')
        self.btn_browse_existing.pack(side=tk.LEFT)
        
        # === URL・個数入力フレーム ===
        input_frame = ttk.LabelFrame(self.root, text='3. 商品情報入力', padding=10)
        input_frame.grid(row=2, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Label(input_frame, text='モノタロウ商品URL:').grid(row=0, column=0, sticky='w')
        self.entry_url = ttk.Entry(input_frame, width=60)
        self.entry_url.grid(row=0, column=1, columnspan=2, sticky='ew', padx=5)
        
        ttk.Label(input_frame, text='個数:').grid(row=1, column=0, sticky='w')
        self.entry_quantity = ttk.Entry(input_frame, width=10)
        self.entry_quantity.grid(row=1, column=1, sticky='w', padx=5)
        self.entry_quantity.insert(0, '1')
        
        ttk.Button(input_frame, text='リストに追加', command=self.add_to_list).grid(row=1, column=2, padx=5)
        
        input_frame.columnconfigure(1, weight=1)
        
        # === リスト表示フレーム ===
        list_frame = ttk.LabelFrame(self.root, text='4. 追加予定の商品リスト', padding=10)
        list_frame.grid(row=3, column=0, columnspan=3, sticky='nsew', padx=10, pady=5)
        
        # スクロールバー付きListbox
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.listbox = tk.Listbox(list_frame, height=8, yscrollcommand=scrollbar.set, font=('Courier', 9))
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # リスト操作ボタン
        button_frame = ttk.Frame(self.root)
        button_frame.grid(row=4, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Button(button_frame, text='選択項目を削除', command=self.remove_from_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text='リストをクリア', command=self.clear_list).pack(side=tk.LEFT, padx=5)
        
        # === 実行ボタン ===
        run_frame = ttk.Frame(self.root)
        run_frame.grid(row=5, column=0, columnspan=3, sticky='ew', padx=10, pady=10)
        
        self.btn_run = ttk.Button(run_frame, text='変換実行', command=self.run_conversion)
        self.btn_run.pack(side=tk.LEFT, padx=5)
        
        # ステータスバー
        self.status_var = tk.StringVar(value='準備完了')
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=6, column=0, columnspan=3, sticky='ew')
        
        self.root.rowconfigure(3, weight=1)
        self.root.columnconfigure(0, weight=1)
        
        self.toggle_mode()
    
    def toggle_mode(self):
        """モード切り替え時のUIの有効/無効を切り替え"""
        if self.mode_var.get() == 'new':
            self.new_file_frame.pack(side=tk.LEFT, padx=10)
            self.append_file_frame.pack_forget()
            self.entry_file_path.config(state='normal')
            self.entry_sheet_name.config(state='normal')
        else:
            self.new_file_frame.pack_forget()
            self.append_file_frame.pack(side=tk.LEFT, padx=10)
            self.entry_existing_file.config(state='normal')
            self.btn_browse_existing.config(state='normal')
    
    def browse_file(self, entry, save=False):
        """ファイルダイアログを開く"""
        if save:
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        else:
            file_path = filedialog.askopenfilename(
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        
        if file_path:
            entry.delete(0, tk.END)
            entry.insert(0, file_path)
    
    def add_to_list(self):
        """URLと個数をリストに追加"""
        url = self.entry_url.get().strip()
        quantity = self.entry_quantity.get().strip()
        
        if not url:
            messagebox.showwarning('警告', 'URLを入力してください')
            return
        
        if not url.startswith('https://www.monotaro.com'):
            messagebox.showwarning('警告', 'モノタロウのURLを入力してください\nhttps://www.monotaro.com/p/...')
            return
        
        if not quantity.isdigit() or int(quantity) < 1:
            messagebox.showwarning('警告', '個数は1以上の数字で入力してください')
            return
        
        self.listbox.insert(tk.END, f'{url} | 個数: {quantity}')
        self.entry_url.delete(0, tk.END)
        self.entry_quantity.delete(0, tk.END)
        self.entry_quantity.insert(0, '1')
        self.entry_url.focus()
    
    def remove_from_list(self):
        """選択した項目をリストから削除"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning('警告', '削除する項目を選択してください')
            return
        self.listbox.delete(selection)
    
    def clear_list(self):
        """リストをすべてクリア"""
        if messagebox.askyesno('確認', 'リスト内容をすべてクリアしますか？'):
            self.listbox.delete(0, tk.END)
    
    def fetch_monotaro_data(self, url):
        """モノタロウの商品ページから商品情報を取得"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return None
            
            # エンコーディング自動検出
            response.encoding = response.apparent_encoding
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 商品コードをURLから取得
            matched = re.search(r'/p/(\d+)/(\d+)', url)
            item_code = None
            if matched:
                item_code = matched.group(1) + matched.group(2)
            
            # 商品名取得
            product_name = ''
            # 複数の可能性をチェック
            name_selectors = [
                '.productName',
                'h1.p-product-name',
                '[data-testid="product-name"]',
                '.p-header h1'
            ]
            for selector in name_selectors:
                tag = soup.select_one(selector)
                if tag:
                    product_name = tag.get_text(strip=True)
                    break
            
            if not product_name:
                return None
            
            # 型番取得
            model_number = ''
            # product-specセクションから型番を探す
            spec_rows = soup.find_all('tr', class_='product-spec')
            for row in spec_rows:
                header = row.find('th')
                if header and '型番' in header.get_text():
                    data = row.find('td')
                    if data:
                        model_number = data.get_text(strip=True)
                        break
            
            # 型番が見つからない場合は別の方法を試す
            if not model_number:
                spec_section = soup.find('section', class_='product-spec')
                if spec_section:
                    spec_items = spec_section.find_all('dl')
                    for spec_item in spec_items:
                        dt = spec_item.find('dt')
                        if dt and '型番' in dt.get_text():
                            dd = spec_item.find('dd')
                            if dd:
                                model_number = dd.get_text(strip=True)
                                break
            
            # 価格取得（税込価格）
            price = ''
            price_selectors = [
                '.p-price',
                '.price',
                '[data-testid="product-price"]',
                '.productPrice'
            ]
            for selector in price_selectors:
                tag = soup.select_one(selector)
                if tag:
                    price_text = tag.get_text(strip=True)
                    # 数字だけ抽出
                    price_match = re.search(r'(\d+(?:,\d+)*)', price_text.replace(',', ''))
                    if price_match:
                        price = price_match.group(1).replace(',', '')
                        break
            
            return {
                'supplier': 'モノタロウ',
                'item_code': item_code or '',
                'product_name': product_name,
                'model_number': model_number or '',
                'price': price or '',
                'url': url
            }
        
        except Exception as e:
            print(f'エラー: {e}')
            return None
    
    def write_to_excel(self, file_path, sheet_name, data_list, append=True):
        """Excelファイルにデータを書き込み"""
        try:
            if append and sheet_name:
                # 既存ファイルに追加
                wb = openpyxl.load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                start_row = ws.max_row + 1
            else:
                # 新規作成
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name if sheet_name else '注文内容'
                start_row = 1
                
                # ヘッダ行を作成
                headers = ['仕入元', '商品コード', '商品名', '型番', '価格(税込)', '個数', 'URL']
                ws.append(headers)
                
                # ヘッダの書式設定
                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                
                start_row = 2
            
            # データを書き込み
            for idx, data in enumerate(data_list):
                ws.append([
                    data.get('supplier', 'モノタロウ'),
                    data.get('item_code', ''),
                    data.get('product_name', ''),
                    data.get('model_number', ''),
                    data.get('price', ''),
                    data.get('quantity', ''),
                    data.get('url', '')
                ])
            
            # 列の幅を自動調整
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
        
        except Exception as e:
            print(f'Excel書き込みエラー: {e}')
            return False
    
    def run_conversion(self):
        """変換を実行"""
        items = self.listbox.get(0, tk.END)
        if not items:
            messagebox.showwarning('警告', 'URLリストに商品を追加してください')
            return
        
        self.btn_run.config(state='disabled')
        self.status_var.set('処理中...')
        self.root.update()
        
        # スレッドで実行（UIをブロックしない）
        thread = threading.Thread(target=self._process_conversion, args=(items,))
        thread.start()
    
    def _process_conversion(self, items):
        """変換処理をスレッドで実行"""
        data_list = []
        total = len(items)
        
        for idx, item in enumerate(items):
            try:
                # URLと個数を抽出
                parts = item.split(' | 個数: ')
                if len(parts) != 2:
                    continue
                
                url = parts[0].strip()
                try:
                    quantity = int(parts[1].strip())
                except:
                    continue
                
                self.status_var.set(f'処理中... ({idx+1}/{total})')
                self.root.update()
                
                # 商品情報を取得
                result = self.fetch_monotaro_data(url)
                if result:
                    result['quantity'] = quantity
                    data_list.append(result)
            
            except Exception as e:
                print(f'処理エラー: {e}')
                continue
        
        # Excelに書き込み
        if not data_list:
            self.root.after(0, lambda: messagebox.showwarning('警告', '有効な商品情報が取得できませんでした'))
            self.btn_run.config(state='normal')
            self.status_var.set('準備完了')
            return
        
        mode = self.mode_var.get()
        try:
            if mode == 'new':
                file_path = self.entry_file_path.get().strip()
                sheet_name = self.entry_sheet_name.get().strip()
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '新規ファイルパスを入力してください'))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                if not sheet_name:
                    sheet_name = '注文内容'
                success = self.write_to_excel(file_path, sheet_name, data_list, append=False)
                message = f'Excelファイルに{len(data_list)}件の商品を書き込みました'
            else:
                file_path = self.entry_existing_file.get().strip()
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '既存ファイルパスを入力してください'))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                # シート名を取得
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_name = wb.sheetnames[0] if wb.sheetnames else '注文内容'
                except:
                    sheet_name = '注文内容'
                success = self.write_to_excel(file_path, sheet_name, data_list, append=True)
                message = f'既存ファイルに{len(data_list)}件の商品を追加しました'
            
            if success:
                self.root.after(0, lambda: messagebox.showinfo('完了', message))
                self.root.after(0, lambda: self.listbox.delete(0, tk.END))
            else:
                self.root.after(0, lambda: messagebox.showerror('エラー', 'Excelファイルの書き込みに失敗しました'))
        
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror('エラー', f'エラーが発生しました: {e}'))
        
        self.root.after(0, lambda: self.btn_run.config(state='normal'))
        self.root.after(0, lambda: self.status_var.set('準備完了'))


if __name__ == '__main__':
    root = tk.Tk()
    app = MonotaroExcelApp(root)
    root.mainloop()
