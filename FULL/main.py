"""
統合注文情報Excel作成ツール
モノタロウ・秋月電子などの商品情報を自動取得してExcelにまとめる
"""
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import threading
import time
import re

# スクレイパーモジュールをインポート
from scraper_monotaro import MonotaroScraper
from scraper_akizuki import AkizukiScraper


class UnifiedOrderApp:
    def __init__(self, root):
        self.root = root
        self.root.title('統合注文情報Excel作成')
        self.root.geometry('900x750')
        
        self.mode_var = tk.StringVar(value='new')
        
        # スクレイパーを登録
        self.scrapers = [
            MonotaroScraper(),
            AkizukiScraper(),
            # 今後他のサイト用スクレイパーをここに追加
        ]
        
        self.current_site = None  # 現在追加中のサイト
        
        self.create_widgets()
    
    def create_widgets(self):
        # === モード選択 ===
        mode_frame = ttk.LabelFrame(self.root, text='1. モード選択')
        mode_frame.pack(fill='x', padx=10, pady=8)
        
        ttk.Radiobutton(mode_frame, text='新規Excel作成', variable=self.mode_var,
                        value='new', command=self.update_mode_state).pack(side='left', padx=5, pady=5)
        ttk.Radiobutton(mode_frame, text='既存Excelに追記', variable=self.mode_var,
                        value='append', command=self.update_mode_state).pack(side='left', padx=5, pady=5)
        
        # === ファイル設定 ===
        file_frame = ttk.LabelFrame(self.root, text='2. ファイル設定')
        file_frame.pack(fill='x', padx=10, pady=8)
        
        # 新規作成
        new_frame = ttk.Frame(file_frame)
        new_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(new_frame, text='保存先:').grid(row=0, column=0, sticky='w')
        self.save_path_var = tk.StringVar()
        self.save_path_entry = ttk.Entry(new_frame, textvariable=self.save_path_var, width=70)
        self.save_path_entry.grid(row=0, column=1, padx=5)
        self.save_path_btn = ttk.Button(new_frame, text='参照', command=self.browse_save_path)
        self.save_path_btn.grid(row=0, column=2, padx=5)
        
        ttk.Label(new_frame, text='シート名:').grid(row=1, column=0, sticky='w', pady=(6, 0))
        self.sheet_name_var = tk.StringVar(value='注文内容')
        self.sheet_name_entry = ttk.Entry(new_frame, textvariable=self.sheet_name_var, width=30)
        self.sheet_name_entry.grid(row=1, column=1, sticky='w', pady=(6, 0))
        
        # 追記
        append_frame = ttk.Frame(file_frame)
        append_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(append_frame, text='既存Excel:').grid(row=0, column=0, sticky='w')
        self.append_file_var = tk.StringVar()
        self.append_file_entry = ttk.Entry(append_frame, textvariable=self.append_file_var, width=70)
        self.append_file_entry.grid(row=0, column=1, padx=5)
        self.append_file_btn = ttk.Button(append_frame, text='参照', command=self.browse_existing_file)
        self.append_file_btn.grid(row=0, column=2, padx=5)
        
        # === 商品入力 ===
        input_frame = ttk.LabelFrame(self.root, text='3. 商品情報入力（商品ページURLと数量）')
        input_frame.pack(fill='x', padx=10, pady=8)
        
        ttk.Label(input_frame, text='商品URL:').grid(row=0, column=0, sticky='w')
        self.url_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.url_var, width=75).grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(input_frame, text='数量:').grid(row=1, column=0, sticky='w')
        self.qty_var = tk.StringVar(value='1')
        ttk.Entry(input_frame, textvariable=self.qty_var, width=10).grid(row=1, column=1, sticky='w', pady=5)
        ttk.Button(input_frame, text='リストに追加', command=self.add_to_list).grid(row=1, column=2, padx=10, pady=5, sticky='w')
        
        # 現在のサイト表示
        self.site_label_var = tk.StringVar(value='現在のサイト: なし')
        ttk.Label(input_frame, textvariable=self.site_label_var, foreground='blue').grid(row=2, column=0, columnspan=3, sticky='w', pady=(5, 0))
        
        # === リスト ===
        list_frame = ttk.LabelFrame(self.root, text='4. 追加予定リスト')
        list_frame.pack(fill='both', expand=True, padx=10, pady=8)
        
        self.listbox = tk.Listbox(list_frame, height=15)
        self.listbox.pack(side='left', fill='both', expand=True, padx=(5, 0), pady=5)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.listbox.yview)
        scrollbar.pack(side='left', fill='y', pady=5)
        self.listbox.config(yscrollcommand=scrollbar.set)
        
        # リスト変更を監視
        self.listbox.bind('<<ListboxSelect>>', self.on_list_change)
        
        btns = ttk.Frame(list_frame)
        btns.pack(side='left', fill='y', padx=8)
        ttk.Button(btns, text='選択削除', command=self.remove_selected).pack(fill='x', pady=4)
        ttk.Button(btns, text='全クリア', command=self.clear_list).pack(fill='x', pady=4)
        
        # === 実行 ===
        run_frame = ttk.LabelFrame(self.root, text='5. 実行')
        run_frame.pack(fill='x', padx=10, pady=8)
        ttk.Button(run_frame, text='変換実行', command=self.run_conversion).pack(side='left', padx=5, pady=5)
        
        # === ステータス ===
        self.status_var = tk.StringVar(value='準備完了')
        status = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', anchor='w')
        status.pack(fill='x', padx=10, pady=(0, 8))
        
        self.update_mode_state()
    
    def update_mode_state(self):
        """モード切り替え時のUI状態更新"""
        mode = self.mode_var.get()
        new_enable = (mode == 'new')
        append_enable = (mode == 'append')
        
        for w, enable in [
            (self.save_path_entry, new_enable),
            (self.save_path_btn, new_enable),
            (self.sheet_name_entry, True),
            (self.append_file_entry, append_enable),
            (self.append_file_btn, append_enable),
        ]:
            try:
                w.config(state=('normal' if enable else 'disabled'))
            except Exception:
                pass
    
    def browse_save_path(self):
        """保存先ファイル選択"""
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel Workbook', '*.xlsx')]
        )
        if path:
            self.save_path_var.set(path)
    
    def browse_existing_file(self):
        """既存ファイル選択"""
        path = filedialog.askopenfilename(
            filetypes=[('Excel Workbook', '*.xlsx')]
        )
        if path:
            self.append_file_var.set(path)
    
    def get_scraper_for_url(self, url: str):
        """URLに対応するスクレイパーを取得"""
        for scraper in self.scrapers:
            if scraper.is_valid_url(url):
                return scraper
        return None
    
    def update_current_site(self):
        """リストの一番上の項目から現在のサイトを判定"""
        if self.listbox.size() == 0:
            self.current_site = None
            self.site_label_var.set('現在のサイト: なし')
            return
        
        # 一番上の項目を取得
        first_item = self.listbox.get(0)
        m = re.match(r'(.+?)\s+\|\s+個数:\s*(\d+)', first_item)
        if not m:
            self.current_site = None
            self.site_label_var.set('現在のサイト: なし')
            return
        
        url = m.group(1).strip()
        scraper = self.get_scraper_for_url(url)
        
        if scraper:
            self.current_site = scraper.get_site_name()
            self.site_label_var.set(f'現在のサイト: {self.current_site}')
        else:
            self.current_site = None
            self.site_label_var.set('現在のサイト: 不明')
    
    def on_list_change(self, event=None):
        """リスト変更時の処理"""
        self.update_current_site()
    
    def add_to_list(self):
        """URLと個数をリストに追加"""
        url = self.url_var.get().strip()
        qty_str = self.qty_var.get().strip()
        
        if not url:
            messagebox.showwarning('警告', '商品URLを入力してください。')
            return
        
        # URLに対応するスクレイパーを取得
        scraper = self.get_scraper_for_url(url)
        if not scraper:
            messagebox.showwarning('警告', '対応していないサイトのURLです。\n対応サイト: モノタロウ、秋月電子通商')
            return
        
        # 現在のサイトと異なる場合は警告
        if self.current_site and scraper.get_site_name() != self.current_site:
            messagebox.showwarning(
                '警告',
                f'現在のリストは「{self.current_site}」の商品です。\n'
                f'「{scraper.get_site_name()}」の商品は追加できません。\n\n'
                f'別のサイトの商品を追加する場合は、リストをクリアしてください。'
            )
            return
        
        if not qty_str.isdigit() or int(qty_str) < 1:
            messagebox.showwarning('警告', '数量は1以上の整数で入力してください。')
            return
        
        item_text = f'{url} | 個数: {qty_str}'
        self.listbox.insert('end', item_text)
        self.url_var.set('')
        self.qty_var.set('1')
        
        # サイト判定を更新
        self.update_current_site()
    
    def remove_selected(self):
        """選択項目を削除"""
        selected = list(self.listbox.curselection())
        for i in reversed(selected):
            self.listbox.delete(i)
        # サイト判定を更新
        self.update_current_site()
    
    def clear_list(self):
        """リストをクリア"""
        self.listbox.delete(0, 'end')
        self.update_current_site()
    
    def set_status(self, text):
        """ステータス更新"""
        self.status_var.set(text)
        self.root.update_idletasks()
    
    def run_conversion(self):
        """変換実行"""
        count = self.listbox.size()
        if count == 0:
            messagebox.showwarning('警告', 'リストに商品がありません。')
            return
        
        mode = self.mode_var.get()
        if mode == 'new':
            if not self.save_path_var.get().strip():
                messagebox.showwarning('警告', '保存先のExcelファイルパスを指定してください。')
                return
            if not self.sheet_name_var.get().strip():
                messagebox.showwarning('警告', 'シート名を入力してください。')
                return
        else:
            if not self.append_file_var.get().strip():
                messagebox.showwarning('警告', '追記先のExcelファイルを指定してください。')
                return
        
        self.set_status('処理を開始します...')
        thread = threading.Thread(target=self.worker_process, daemon=True)
        thread.start()
    
    def worker_process(self):
        """ワーカースレッド - 商品情報取得とExcel書き込み"""
        try:
            items = []
            failed_urls = []  # 失敗したURL
            
            for i in range(self.listbox.size()):
                text = self.listbox.get(i)
                m = re.match(r'(.+?)\s+\|\s+個数:\s*(\d+)', text)
                if not m:
                    continue
                
                url = m.group(1).strip()
                qty = int(m.group(2))
                
                self.set_status(f'取得中 ({i+1}/{self.listbox.size()}): {url[:50]}...')
                
                # URLに対応するスクレイパーを取得
                scraper = self.get_scraper_for_url(url)
                if not scraper:
                    failed_urls.append(url)
                    continue
                
                # 商品情報を取得
                product = scraper.fetch_product_data(url)
                if product is None:
                    failed_urls.append(url)
                    # 1件でも失敗したら即エラー表示・処理中断（モノタロウ仕様）
                    error_message = '以下のURLを正しく開けませんでした。\nURLが商品ページまで指定していることを確認してください:\n\n'
                    error_message += f'• {url}\n'
                    self.root.after(0, lambda msg=error_message: messagebox.showerror('エラー', msg))
                    self.set_status('準備完了')
                    return
                
                product['quantity'] = qty
                items.append(product)
                
                # サイトに応じた待機時間
                if scraper.get_site_name() == '秋月電子通商':
                    time.sleep(0.7)
                else:
                    time.sleep(0.1)
            
            if not items:
                messagebox.showwarning('警告', '情報を取得できませんでした。ページ構造の変更やメンテナンスの可能性があります。')
                self.set_status('準備完了')
                return
            
            # Excel書き込み
            mode = self.mode_var.get()
            if mode == 'new':
                self.write_new_excel(
                    self.save_path_var.get().strip(),
                    self.sheet_name_var.get().strip(),
                    items
                )
                messagebox.showinfo('完了', f'Excelを作成しました:\n{self.save_path_var.get().strip()}')
            else:
                sheet = self.sheet_name_var.get().strip() or '注文内容'
                self.append_to_excel(
                    self.append_file_var.get().strip(),
                    sheet,
                    items
                )
                messagebox.showinfo('完了', f'Excelに追記しました:\n{self.append_file_var.get().strip()}')
            
            self.clear_list()
            self.set_status('準備完了')
        
        except PermissionError as pe:
            self.root.after(0, lambda msg=str(pe): messagebox.showerror('エラー', msg))
            self.set_status('準備完了')
        except Exception as e:
            messagebox.showerror('エラー', f'処理中にエラーが発生しました:\n{e}')
            self.set_status('準備完了')
    
    def write_new_excel(self, path: str, sheet_name: str, items: list):
        """新規Excel作成"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        headers = ['メーカー', '注文コード', '商品名', '品番/型番', '単価（税別）', '数量', '値段（税別）', 'URL', '税込み']
        ws.append(headers)
        
        # ヘッダ書式設定（秋月スタイル）
        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='DDDDDD')
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データ行
        for item in items:
            price_ex = item.get('price_excl_tax', 0) or 0
            qty = item.get('quantity', 0) or 0
            
            # 価格を数値に変換
            if isinstance(price_ex, str):
                try:
                    price_ex = int(price_ex.replace(',', ''))
                except:
                    price_ex = 0
            
            total_ex = price_ex * qty
            
            price_in = item.get('price_incl_tax', 0) or 0
            if isinstance(price_in, str):
                try:
                    price_in = int(price_in.replace(',', ''))
                except:
                    price_in = 0
            
            row = [
                item.get('supplier', ''),
                str(item.get('item_code', '')),
                item.get('name', ''),
                str(item.get('model', '')),
                price_ex,
                qty,
                total_ex,
                item.get('url', ''),
                price_in,
            ]
            ws.append(row)
            
            # 商品コードと型番を文字列として設定
            current_row = ws.max_row
            try:
                cell_code = ws.cell(row=current_row, column=2)
                cell_code.value = str(item.get('item_code', ''))
                cell_code.number_format = '@'
            except:
                pass
            
            try:
                cell_model = ws.cell(row=current_row, column=4)
                cell_model.value = str(item.get('model', ''))
                cell_model.number_format = '@'
            except:
                pass
        
        # 列幅自動調整（秋月スタイル）
        self.autofit_columns(ws, max_width=50)
        
        try:
            wb.save(path)
        except PermissionError:
            raise PermissionError(f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{path}')
    
    def append_to_excel(self, path: str, sheet_name: str, items: list):
        """既存Excelに追記"""
        try:
            wb = openpyxl.load_workbook(path)
        except PermissionError:
            raise PermissionError(f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{path}')
        except FileNotFoundError:
            # ファイルが存在しない場合は新規作成
            self.write_new_excel(path, sheet_name, items)
            return
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            headers = ['メーカー', '注文コード', '商品名', '品番/型番', '単価（税別）', '数量', '値段（税別）', 'URL', '税込み']
            ws.append(headers)
            header_font = Font(bold=True)
            header_fill = PatternFill('solid', fgColor='DDDDDD')
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データ行
        for item in items:
            price_ex = item.get('price_excl_tax', 0) or 0
            qty = item.get('quantity', 0) or 0
            
            if isinstance(price_ex, str):
                try:
                    price_ex = int(price_ex.replace(',', ''))
                except:
                    price_ex = 0
            
            total_ex = price_ex * qty
            
            price_in = item.get('price_incl_tax', 0) or 0
            if isinstance(price_in, str):
                try:
                    price_in = int(price_in.replace(',', ''))
                except:
                    price_in = 0
            
            row = [
                item.get('supplier', ''),
                str(item.get('item_code', '')),
                item.get('name', ''),
                str(item.get('model', '')),
                price_ex,
                qty,
                total_ex,
                item.get('url', ''),
                price_in,
            ]
            ws.append(row)
            
            # 商品コードと型番を文字列として設定
            current_row = ws.max_row
            try:
                cell_code = ws.cell(row=current_row, column=2)
                cell_code.value = str(item.get('item_code', ''))
                cell_code.number_format = '@'
            except:
                pass
            
            try:
                cell_model = ws.cell(row=current_row, column=4)
                cell_model.value = str(item.get('model', ''))
                cell_model.number_format = '@'
            except:
                pass
        
        # 列幅自動調整
        self.autofit_columns(ws, max_width=50)
        
        try:
            wb.save(path)
        except PermissionError:
            raise PermissionError(f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{path}')
    
    def autofit_columns(self, ws, max_width=50):
        """列幅を自動調整（秋月スタイル）"""
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for idx, val in enumerate(row, start=1):
                length = len(str(val)) if val is not None else 0
                widths[idx] = max(widths.get(idx, 0), length)
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 10), max_width)


def main():
    root = tk.Tk()
    app = UnifiedOrderApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
