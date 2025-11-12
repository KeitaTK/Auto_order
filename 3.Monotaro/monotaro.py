import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
import requests
from bs4 import BeautifulSoup
import re
import threading
import time

class MonotaroExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title('モノタロウ注文情報Excel作成')
        self.root.geometry('900x750')
        
        self.mode_var = tk.StringVar(value='new')
        self.data_list = []
        
        # セッション設定（ボット対策対応）
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ja-JP,ja;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })
        
        self.create_widgets()
        
    def create_widgets(self):
        # === モード選択フレーム ===
        mode_frame = ttk.LabelFrame(self.root, text='1. モード選択', padding=10)
        mode_frame.grid(row=0, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Radiobutton(mode_frame, text='新規作成', variable=self.mode_var, 
                       value='new', command=self.toggle_mode).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(mode_frame, text='既存ファイルに追加', variable=self.mode_var, 
                       value='append', command=self.toggle_mode).pack(side=tk.LEFT, padx=10)
        
        # === ファイル設定フレーム ===
        file_frame = ttk.LabelFrame(self.root, text='2. ファイル設定', padding=10)
        file_frame.grid(row=1, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        # 新規作成用
        self.new_file_frame = ttk.Frame(file_frame)
        self.new_file_frame.pack(side=tk.LEFT, padx=10)
        
        ttk.Label(self.new_file_frame, text='ファイルパス:').pack()
        file_path_frame = ttk.Frame(self.new_file_frame)
        file_path_frame.pack(pady=5)
        self.entry_file_path = ttk.Entry(file_path_frame, width=40)
        self.entry_file_path.pack(side=tk.LEFT, padx=5)
        ttk.Button(file_path_frame, text='参照', 
                  command=lambda: self.browse_file(self.entry_file_path, save=True)).pack(side=tk.LEFT)
        
        ttk.Label(self.new_file_frame, text='シート名:').pack(pady=(10, 0))
        self.entry_sheet_name = ttk.Entry(self.new_file_frame, width=20)
        self.entry_sheet_name.pack(pady=5)
        self.entry_sheet_name.insert(0, '注文内容')
        
        # 既存ファイル用
        self.append_file_frame = ttk.Frame(file_frame)
        
        ttk.Label(self.append_file_frame, text='ファイルパス:').pack()
        append_path_frame = ttk.Frame(self.append_file_frame)
        append_path_frame.pack(pady=5)
        self.entry_existing_file = ttk.Entry(append_path_frame, width=40, state='disabled')
        self.entry_existing_file.pack(side=tk.LEFT, padx=5)
        self.btn_browse_existing = ttk.Button(append_path_frame, text='参照', 
                                             command=lambda: self.browse_file(self.entry_existing_file, save=False),
                                             state='disabled')
        self.btn_browse_existing.pack(side=tk.LEFT)
        
        # === URL・個数入力フレーム ===
        input_frame = ttk.LabelFrame(self.root, text='3. 商品情報入力', padding=10)
        input_frame.grid(row=2, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Label(input_frame, text='モノタロウ商品URL:').grid(row=0, column=0, sticky='w')
        self.entry_url = ttk.Entry(input_frame, width=60)
        self.entry_url.grid(row=0, column=1, columnspan=2, sticky='ew', padx=5)
        
        ttk.Label(input_frame, text='個数:').grid(row=1, column=0, sticky='w')
        self.entry_quantity = ttk.Entry(input_frame, width=10)
        self.entry_quantity.grid(row=1, column=1, sticky='w', padx=5)
        self.entry_quantity.insert(0, '1')
        
        ttk.Button(input_frame, text='リストに追加', command=self.add_to_list).grid(row=1, column=2, padx=5)
        
        input_frame.columnconfigure(1, weight=1)
        
        # === リスト表示フレーム ===
        list_frame = ttk.LabelFrame(self.root, text='4. 追加予定の商品リスト', padding=10)
        list_frame.grid(row=3, column=0, columnspan=3, sticky='nsew', padx=10, pady=5)
        
        # スクロールバー付きListbox
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.listbox = tk.Listbox(list_frame, height=8, yscrollcommand=scrollbar.set, font=('Courier', 9))
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # リスト操作ボタン
        button_frame = ttk.Frame(self.root)
        button_frame.grid(row=4, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        
        ttk.Button(button_frame, text='選択項目を削除', command=self.remove_from_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text='リストをクリア', command=self.clear_list).pack(side=tk.LEFT, padx=5)
        
        # === 実行ボタン ===
        run_frame = ttk.Frame(self.root)
        run_frame.grid(row=5, column=0, columnspan=3, sticky='ew', padx=10, pady=10)
        
        self.btn_run = ttk.Button(run_frame, text='変換実行', command=self.run_conversion)
        self.btn_run.pack(side=tk.LEFT, padx=5)
        
        # ステータスバー
        self.status_var = tk.StringVar(value='準備完了')
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=6, column=0, columnspan=3, sticky='ew')
        
        self.root.rowconfigure(3, weight=1)
        self.root.columnconfigure(0, weight=1)
        
        self.toggle_mode()
    
    def toggle_mode(self):
        """モード切り替え時のUIの有効/無効を切り替え"""
        if self.mode_var.get() == 'new':
            self.new_file_frame.pack(side=tk.LEFT, padx=10)
            self.append_file_frame.pack_forget()
            self.entry_file_path.config(state='normal')
            self.entry_sheet_name.config(state='normal')
        else:
            self.new_file_frame.pack_forget()
            self.append_file_frame.pack(side=tk.LEFT, padx=10)
            self.entry_existing_file.config(state='normal')
            self.btn_browse_existing.config(state='normal')
    
    def browse_file(self, entry, save=False):
        """ファイルダイアログを開く"""
        if save:
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        else:
            file_path = filedialog.askopenfilename(
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        
        if file_path:
            entry.delete(0, tk.END)
            entry.insert(0, file_path)
    
    def add_to_list(self):
        """URLと個数をリストに追加"""
        url = self.entry_url.get().strip()
        quantity = self.entry_quantity.get().strip()
        
        if not url:
            messagebox.showwarning('警告', 'URLを入力してください')
            return
        
        if not url.startswith('https://www.monotaro.com'):
            messagebox.showwarning('警告', 'モノタロウのURLを入力してください\nhttps://www.monotaro.com/p/...')
            return
        
        if not quantity.isdigit() or int(quantity) < 1:
            messagebox.showwarning('警告', '個数は1以上の数字で入力してください')
            return
        
        self.listbox.insert(tk.END, f'{url} | 個数: {quantity}')
        self.entry_url.delete(0, tk.END)
        self.entry_quantity.delete(0, tk.END)
        self.entry_quantity.insert(0, '1')
        self.entry_url.focus()
    
    def remove_from_list(self):
        """選択した項目をリストから削除"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning('警告', '削除する項目を選択してください')
            return
        self.listbox.delete(selection)
    
    def clear_list(self):
        """リストをすべてクリア"""
        if messagebox.askyesno('確認', 'リスト内容をすべてクリアしますか？'):
            self.listbox.delete(0, tk.END)
    
    def fetch_monotaro_data(self, url):
        """モノタロウの商品ページから商品情報を取得"""
        try:
            # リトライロジック
            response = None
            for attempt in range(3):
                try:
                    response = self.session.get(url, timeout=15)
                    # ステータスコード確認
                    if response.status_code == 403 or 'ログイン' in response.text:
                        # ボット対策またはログイン要求
                        if attempt < 2:
                            time.sleep(2)  # 待機してリトライ
                            continue
                        else:
                            print(f'ログイン要求またはボット対策: {url}')
                            return None
                    if response.status_code != 200:
                        print(f'HTTPエラー {response.status_code}: {url}')
                        return None
                    break
                except requests.exceptions.Timeout:
                    if attempt < 2:
                        time.sleep(2)
                        continue
                    raise
            if response is None:
                print(f'HTTPレスポンスが取得できませんでした: {url}')
                return None
            # エンコーディング自動検出
            response.encoding = response.apparent_encoding or 'utf-8'
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # URLからコード抽出
            matched = re.search(r'/p/(\d+)/(\d+)', url)
            item_code = None
            if matched:
                item_code = matched.group(1) + matched.group(2)
            
            # ============ 商品名取得 ============
            product_name = ''
            
            # h1タグから取得（メイン商品名）
            h1_tag = soup.select_one('h1')
            if h1_tag:
                product_name = h1_tag.get_text(strip=True)
            
            # h1が見つからない場合は別の方法
            if not product_name:
                title_tag = soup.select_one('title')
                if title_tag:
                    title_text = title_tag.get_text()
                    # タイトルから最初の部分を抽出（"商品名 | モノタロウ"形式）
                    if '|' in title_text:
                        product_name = title_text.split('|')[0].strip()
                    else:
                        product_name = title_text.strip()
            
            if not product_name:
                print(f'商品名が見つかりません: {url}')
                return None
            
            # ============ 型番取得 ============
            model_number = ''
            
            # 方法1: span.AttributeLabelItem から取得（品番M2.5×16 形式）
            attr_labels = soup.select('span.AttributeLabelItem')
            for attr_label in attr_labels:
                text = attr_label.get_text(strip=True)
                # "品番M2.5×16" から "M2.5×16" を抽出
                if '品番' in text:
                    match = re.search(r'品番(.+)', text)
                    if match:
                        model_number = match.group(1).strip()
                        break
            
            # 方法2: タイトルから型番を抽出（M2.5×16 商品名... 形式）
            if not model_number:
                title_tag = soup.select_one('title')
                if title_tag:
                    title_text = title_tag.get_text(strip=True)
                    # パターン: 型番が先頭にある（M2.5×16のような形式）
                    # 数字、アルファベット、×、-、.などを含む可能性
                    match = re.match(r'^([A-Z0-9\.\-]+[×x][A-Z0-9\.\-]+)\s+', title_text, re.IGNORECASE)
                    if match:
                        model_number = match.group(1)
            
            # 方法3: dl > dt/dd から取得（構造化データ）
            if not model_number:
                dts = soup.find_all('dt')
                for dt in dts:
                    dt_text = dt.get_text(strip=True)
                    if '型番' in dt_text or 'SKU' in dt_text or '品番' in dt_text:
                        dd = dt.find_next('dd')
                        if dd:
                            model_number = dd.get_text(strip=True)
                            break
            
            # 方法4: tableのtdから取得
            if not model_number:
                rows = soup.find_all('tr')
                for row in rows:
                    cells = row.find_all(['th', 'td'])
                    if cells and len(cells) >= 2:
                        header = cells[0].get_text(strip=True)
                        if '型番' in header or '品番' in header:
                            model_number = cells[1].get_text(strip=True)
                            break
            
            # ============ 価格取得 ============
            price_tax_excluded = ''  # 税別価格
            price_tax_included = ''  # 税込価格
            
            # 方法1: 販売価格(税別)と販売価格(税込)を直接取得
            # 税別価格: SellingPrice__Title の次の Price--Lg
            selling_price_title = soup.select_one('.SellingPrice__Title')
            if selling_price_title:
                price_elem = selling_price_title.find_next('span', class_='Price--Lg')
                if price_elem:
                    price_text = price_elem.get_text(strip=True)
                    numbers = re.findall(r'\d+', price_text.replace(',', ''))
                    if numbers:
                        price_tax_excluded = numbers[0]
            
            # 税込価格: 販売価格(税込)を含むReferencePrice
            ref_price_title = soup.find('span', class_='ReferencePrice__Title', text=re.compile(r'販売価格.*税込'))
            if ref_price_title and ref_price_title.parent:
                parent_text = ref_price_title.parent.get_text(strip=True)
                numbers = re.findall(r'[\d,]+', parent_text)
                if numbers:
                    price_tax_included = numbers[-1].replace(',', '')
            
            # 方法2: フォールバック - 複数の価格セレクタを試す（税込価格として扱う）
            if not price_tax_included and not price_tax_excluded:
                price_patterns = [
                    'span[class*="price"]',
                    'span[class*="Price"]',
                    '.p-price',
                    '.productPrice',
                    'div[data-testid*="price"]'
                ]
                
                for pattern in price_patterns:
                    price_elements = soup.select(pattern)
                    for elem in price_elements:
                        price_text = elem.get_text(strip=True)
                        # 数字を抽出
                        numbers = re.findall(r'\d+', price_text.replace(',', ''))
                        if numbers:
                            # 最も大きい数字（通常は価格）を税込として扱う
                            price_tax_included = max(numbers, key=int)
                            break
                    if price_tax_included:
                        break
            
            # 方法3: 価格が見つからない場合、テキスト内を直接探索
            if not price_tax_included and not price_tax_excluded:
                page_text = soup.get_text()
                # 「¥xxx」パターンで検索
                price_match = re.search(r'¥([\d,]+)', page_text)
                if price_match:
                    price_tax_included = price_match.group(1).replace(',', '')
            
            # 税別が取得できて税込がない場合は計算
            if price_tax_excluded and not price_tax_included:
                try:
                    price_tax_included = str(int(float(price_tax_excluded) * 1.1))
                except:
                    pass
            
            # 税込が取得できて税別がない場合は計算
            if price_tax_included and not price_tax_excluded:
                try:
                    price_tax_excluded = str(int(float(price_tax_included) / 1.1))
                except:
                    pass
            
            return {
                'supplier': 'モノタロウ',
                'item_code': item_code or '',
                'product_name': product_name,
                'model_number': model_number or '',
                'price_tax_excluded': price_tax_excluded or '',
                'price_tax_included': price_tax_included or '',
                'url': url
            }
        
        except Exception as e:
            print(f'エラー: {str(e)}, URL: {url}')
            return None
    
    def write_to_excel(self, file_path, sheet_name, data_list, append=True):
        """Excelファイルにデータを書き込み"""
        try:
            wb = None
            ws: Worksheet = None  # 型ヒント追加
            if append and sheet_name:
                # 既存ファイルに追加
                try:
                    wb = openpyxl.load_workbook(file_path)
                except PermissionError:
                    raise PermissionError(f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{file_path}')
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                start_row = ws.max_row + 1
            else:
                # 新規作成
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name if sheet_name else '注文内容'
                # ヘッダ行を作成
                headers = ['メーカー', '注文コード', '商品名', '品番/型番', '単価', '数量', '値段（税別）', 'URL', '税込み']
                ws.append(headers)
                # ヘッダの書式設定
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                start_row = 2
            
            # データを書き込み
            for idx, data in enumerate(data_list):
                price_tax_excluded = data.get('price_tax_excluded', '')
                price_tax_included = data.get('price_tax_included', '')
                quantity = data.get('quantity', '')
                total_tax_excluded = ''
                if price_tax_excluded and quantity:
                    try:
                        price_num = float(str(price_tax_excluded).replace(',', ''))
                        qty_num = int(str(quantity))
                        total_tax_excluded = str(int(price_num * qty_num))
                    except:
                        pass
                item_code = data.get('item_code', '')
                model_number = data.get('model_number', '')
                row = [
                    data.get('supplier', 'モノタロウ'),  # メーカー
                    item_code,  # 注文コード
                    data.get('product_name', ''),  # 商品名
                    model_number,  # 品番/型番
                    price_tax_excluded,  # 単価
                    quantity,  # 数量
                    total_tax_excluded,  # 値段（税別）
                    data.get('url', ''),  # URL
                    price_tax_included  # 税込み
                ]
                ws.append(row)
                current_row = ws.max_row
                # 商品コード（B列=2）
                if item_code:
                    try:
                        cell_item_code = ws.cell(row=current_row, column=2)
                        cell_item_code.value = str(item_code)
                        cell_item_code.number_format = '@'
                    except Exception:
                        pass
                # 型番（D列=4）
                if model_number:
                    try:
                        cell_model = ws.cell(row=current_row, column=4)
                        cell_model.value = str(model_number)
                        cell_model.number_format = '@'
                    except Exception:
                        pass
            
            # 列の幅を自動調整
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            try:
                wb.save(file_path)
            except PermissionError:
                raise PermissionError(f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{file_path}')
            return True
        
        except PermissionError as pe:
            # PermissionErrorは再発生させて上位で処理
            raise pe
        except Exception as e:
            print(f'Excel書き込みエラー: {e}')
            return False
    
    def run_conversion(self):
        """変換を実行"""
        items = self.listbox.get(0, tk.END)
        if not items:
            messagebox.showwarning('警告', 'URLリストに商品を追加してください')
            return
        
        self.btn_run.config(state='disabled')
        self.status_var.set('処理中...')
        self.root.update()
        
        # スレッドで実行（UIをブロックしない）
        thread = threading.Thread(target=self._process_conversion, args=(items,))
        thread.daemon = True
        thread.start()
    
    def _process_conversion(self, items):
        """変換処理をスレッドで実行"""
        data_list = []
        total = len(items)
        success_count = 0
        
        for idx, item in enumerate(items):
            try:
                # URLと個数を抽出
                parts = item.split(' | 個数: ')
                if len(parts) != 2:
                    continue
                url = parts[0].strip()
                try:
                    quantity = int(parts[1].strip())
                except:
                    continue
                self.status_var.set(f'処理中... ({idx+1}/{total})')
                self.root.update()
                # 商品情報を取得
                result = self.fetch_monotaro_data(url)
                if result:
                    result['quantity'] = quantity
                    data_list.append(result)
                    success_count += 1
                else:
                    # 1件でも失敗したら即エラー表示・処理中断
                    error_message = '以下のURLを正しく開けませんでした。\nURLが商品ページまで指定していることを確認してください:\n\n'
                    error_message += f'• {url}\n'
                    self.root.after(0, lambda msg=error_message: messagebox.showerror('エラー', msg))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                # レート制限対策（サーバーに優しい）
                time.sleep(0.1)
            except Exception as e:
                print(f'処理エラー: {e}')
                continue
        
        # Excelに書き込み
        if not data_list:
            self.root.after(0, lambda: messagebox.showwarning('警告', '有効な商品情報が取得できませんでした\n（ページ構造が変わった可能性があります）'))
            self.btn_run.config(state='normal')
            self.status_var.set('準備完了')
            return
        
        mode = self.mode_var.get()
        try:
            if mode == 'new':
                file_path = self.entry_file_path.get().strip()
                sheet_name = self.entry_sheet_name.get().strip()
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '新規ファイルパスを入力してください'))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                if not sheet_name:
                    sheet_name = '注文内容'
                success = self.write_to_excel(file_path, sheet_name, data_list, append=False)
                message = f'Excelファイルに{len(data_list)}件の商品を書き込みました'
            else:
                file_path = self.entry_existing_file.get().strip()
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '既存ファイルパスを入力してください'))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                # シート名を取得
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_name = wb.sheetnames[0] if wb.sheetnames else '注文内容'
                except PermissionError:
                    self.root.after(0, lambda fp=file_path: messagebox.showerror(
                        'エラー', 
                        f'ファイルが他のアプリケーションで開かれています。\nファイルを閉じてから実行してください:\n{fp}'
                    ))
                    self.btn_run.config(state='normal')
                    self.status_var.set('準備完了')
                    return
                except:
                    sheet_name = '注文内容'
                success = self.write_to_excel(file_path, sheet_name, data_list, append=True)
                message = f'既存ファイルに{len(data_list)}件の商品を追加しました'
            
            if success:
                self.root.after(0, lambda: messagebox.showinfo('完了', message))
                self.root.after(0, lambda: self.listbox.delete(0, tk.END))
            else:
                self.root.after(0, lambda: messagebox.showerror('エラー', 'Excelファイルの書き込みに失敗しました'))
        
        except PermissionError as pe:
            self.root.after(0, lambda msg=str(pe): messagebox.showerror('エラー', msg))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror('エラー', f'エラーが発生しました: {e}'))
        
        self.root.after(0, lambda: self.btn_run.config(state='normal'))
        self.root.after(0, lambda: self.status_var.set('準備完了'))


if __name__ == '__main__':
    root = tk.Tk()
    app = MonotaroExcelApp(root)
    root.mainloop()
