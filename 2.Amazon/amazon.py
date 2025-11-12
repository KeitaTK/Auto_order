# -*- coding: utf-8 -*-
# Amazon.co.jp 商品URL＋数量 → 商品名・型番・税込/税別価格の取得とExcel出力（Tkinter GUI）
# 対策版（exe対応）:
#   - multiprocessing.freeze_support() でPyInstallerの凍結環境に対応（自己再起動防止）
#   - Selenium(undetected-chromedriver)は use_subprocess=True で安定化
#   - QuietChrome (__del__) で WinError 6 の終了時例外を抑止
#   - HTTPフォールバック、指数バックオフ、CAPTCHA手動解決、Ctrl+C穏当終了を維持

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import random
import re
import os
import sys
import shutil
import platform

import requests
from bs4 import BeautifulSoup

import openpyxl
from openpyxl.utils import get_column_letter

# Selenium / undetected-chromedriver
try:
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException
    SEL_AVAILABLE = True
except Exception:
    SEL_AVAILABLE = False

# WinError 6 を抑止するため、__del__での不要エラーを握るサブクラス
if SEL_AVAILABLE:
    class QuietChrome(uc.Chrome):
        def __del__(self):
            try:
                try:
                    self.service.process.kill()
                except Exception:
                    pass
                try:
                    super().quit()
                except Exception:
                    pass
            except Exception:
                pass

UA_POOL = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
    '(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
]

TAX_RATE = 0.10  # 日本の消費税 10%

def parse_price_to_int(text: str) -> int:
    if not text:
        return 0
    num = re.sub(r'[^\d]', '', text)
    try:
        return int(num)
    except:
        return 0

def safe_text(node):
    return node.get_text(strip=True) if node else ''

def amazon_like_headers():
    return {
        'User-Agent': random.choice(UA_POOL),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'ja-JP,ja;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://www.amazon.co.jp/'
    }

def is_amazon_product_url(url: str) -> bool:
    if not url.startswith('https://www.amazon.co.jp/'):
        return False
    return ('/dp/' in url) or ('/gp/product/' in url)

def extract_asin_from_url(url: str) -> str:
    m = re.search(r'/dp/([A-Z0-9]{10})', url)
    if m:
        return m.group(1)
    m = re.search(r'/gp/product/([A-Z0-9]{10})', url)
    if m:
        return m.group(1)
    return ''

def extract_title(soup: BeautifulSoup) -> str:
    t = soup.select_one('#productTitle')
    if t:
        return safe_text(t)
    tt = soup.find('title')
    if tt and tt.get_text():
        return tt.get_text().split(': Amazon')[0].strip()
    return ''

def extract_price_tax_included(soup: BeautifulSoup) -> int:
    candidates = [
        '#corePriceDisplay_desktop_feature_div .a-price .a-offscreen',
        '#apex_desktop .a-price .a-offscreen',
        'span.a-price .a-offscreen',
        '#priceblock_ourprice',
        '#priceblock_dealprice',
        '#priceblock_saleprice',
        'span#sns-base-price',
        '.apexPriceToPay .a-offscreen',
    ]
    for sel in candidates:
        for n in soup.select(sel):
            p = parse_price_to_int(n.get_text())
            if p > 0:
                return p
    m = re.search(r'[¥￥]\s?([\d,]+)', soup.get_text())
    if m:
        return parse_price_to_int(m.group(0))
    return 0

def extract_model_number_and_asin_from_tables(soup: BeautifulSoup):
    model = ''
    asin = ''
    keys = ['型番', 'モデル番号', '品番', 'Item model number', 'Manufacturer Part Number', 'メーカー型番', 'ASIN']

    table_selectors = [
        '#productDetails_techSpec_section_1 tr',
        '#productDetails_detailBullets_sections1 tr',
        '#productDetails_db_sections tr'
    ]
    for sel in table_selectors:
        for tr in soup.select(sel):
            th = tr.find('th')
            td = tr.find('td')
            if not th or not td:
                continue
            label = safe_text(th)
            value = safe_text(td)
            if not label:
                continue
            if 'ASIN' in label and not asin and value:
                asin = value.strip()
            elif any(k in label for k in keys if k != 'ASIN'):
                if not model and value:
                    model = value.strip()

    if not model or not asin:
        for li in soup.select('#detailBullets_feature_div li'):
            bold = li.find('span', class_='a-text-bold')
            if not bold:
                continue
            label = bold.get_text(strip=True).rstrip(':：')
            vals = []
            for span in li.find_all('span'):
                if span is not bold:
                    vals.append(span.get_text(' ', strip=True))
            value = ' '.join(vals).strip()
            if not value:
                continue
            if 'ASIN' in label and not asin:
                asin = value
            elif any(k in label for k in keys if k != 'ASIN') and not model:
                model = value

    return model, asin

def compute_tax_pair(price_incl: int, price_excl: int):
    if price_incl > 0 and price_excl == 0:
        price_excl = int(round(price_incl / (1 + TAX_RATE)))
    elif price_excl > 0 and price_incl == 0:
        price_incl = int(round(price_excl * (1 + TAX_RATE)))
    return price_incl, price_excl

def requests_fetch_amazon(url: str, session: requests.Session):
    backoff = 2.0
    for attempt in range(5):
        try:
            resp = session.get(url, timeout=25)
            text_l = resp.text.lower()
            if resp.status_code in (429, 503) or 'captcha' in text_l or 'robot check' in text_l or 'validatecaptcha' in text_l:
                if attempt < 4:
                    sleep_s = backoff + random.uniform(0.5, 1.8)
                    time.sleep(sleep_s)
                    backoff *= 2
                    continue
                return None, 'bot_block'
            if resp.status_code != 200:
                if attempt < 4:
                    sleep_s = backoff + random.uniform(0.5, 1.8)
                    time.sleep(sleep_s)
                    backoff *= 2
                    continue
                return None, f'http_{resp.status_code}'
            break
        except requests.exceptions.RequestException:
            if attempt < 4:
                sleep_s = backoff + random.uniform(0.5, 1.8)
                time.sleep(sleep_s)
                backoff *= 2
                continue
            return None, 'network'

    resp.encoding = resp.apparent_encoding or 'utf-8'
    soup = BeautifulSoup(resp.content, 'html.parser')

    asin = extract_asin_from_url(url)
    title = extract_title(soup)
    model, asin_from_page = extract_model_number_and_asin_from_tables(soup)
    if not asin and asin_from_page:
        asin = asin_from_page.strip()

    price_incl = extract_price_tax_included(soup)
    price_excl = 0
    price_incl, price_excl = compute_tax_pair(price_incl, price_excl)

    if not title and not price_incl and not asin:
        return None, 'no_data'

    return {
        'supplier': 'Amazon',
        'item_code': asin or '',
        'product_name': title or '',
        'model_number': model or '',
        'price_tax_excluded': str(price_excl) if price_excl else '',
        'price_tax_included': str(price_incl) if price_incl else '',
        'url': url
    }, None

def detect_chrome_binary() -> str:
    # 任意指定: 環境変数 CHROME_BINARY があれば最優先
    env_path = os.environ.get('CHROME_BINARY') or os.environ.get('GOOGLE_CHROME_BIN')
    if env_path and os.path.exists(env_path):
        return env_path
    # Windows の既定インストール例
    if platform.system().lower().startswith('win'):
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        # PATH 上のchrome/msedge
        for name in ('chrome.exe', 'msedge.exe'):
            found = shutil.which(name)
            if found:
                return found
    else:
        # macOS / Linux 想定
        for name in ('google-chrome', 'chromium', 'chrome', 'msedge', 'chromium-browser'):
            found = shutil.which(name)
            if found:
                return found
    return ''  # 見つからなければ空

def build_driver(headless=True, proxy=None):
    if not SEL_AVAILABLE:
        raise RuntimeError('Selenium/undetected_chromedriver が利用できません。pipで導入してください。')

    ua = random.choice(UA_POOL)
    options = uc.ChromeOptions()
    if headless:
        options.add_argument('--headless=new')
    options.add_argument(f'--user-agent={ua}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1200,900')
    if proxy:
        options.add_argument(f'--proxy-server={proxy}')

    # 凍結環境での自己再起動問題緩和のため、ブラウザ実行ファイルの明示指定を試みる
    bin_path = detect_chrome_binary()
    if bin_path:
        options.binary_location = bin_path

    # exe凍結環境では use_subprocess=True の方が安定する報告が多い
    driver = QuietChrome(options=options, use_subprocess=True)
    return driver

def selenium_fetch_amazon(driver, url: str, manual_captcha=True, wait_timeout=20):
    backoff = 2.0
    for attempt in range(4):
        try:
            driver.get(url)
        except WebDriverException:
            if attempt < 3:
                time.sleep(backoff + random.uniform(0.4, 1.4))
                backoff *= 2
                continue
            return None, 'driver_get'

        current_url = driver.current_url.lower()
        page_src = driver.page_source.lower()

        if ('validatecaptcha' in current_url or
            'captcha' in page_src or
            'robot check' in page_src):
            if manual_captcha and '--headless' not in ' '.join(driver.capabilities.get('chrome', {}).get('args', [])):
                messagebox.showinfo('確認', 'CAPTCHAが検出されました。\nブラウザで手動解決後にOKを押してください。')
                solved = False
                start = time.time()
                while time.time() - start < 120:
                    cu = driver.current_url.lower()
                    if 'validatecaptcha' not in cu and 'captcha' not in driver.page_source.lower():
                        solved = True
                        break
                    time.sleep(1.5)
                if not solved:
                    return None, 'captcha_timeout'
            else:
                if attempt < 3:
                    time.sleep(backoff + random.uniform(0.8, 1.8))
                    backoff *= 2
                    continue
                return None, 'bot_block'

        try:
            WebDriverWait(driver, wait_timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#productTitle'))
            )
        except TimeoutException:
            if attempt < 3:
                time.sleep(backoff + random.uniform(0.5, 1.5))
                backoff *= 2
                continue
            return None, 'timeout'

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        asin = extract_asin_from_url(url)
        title = extract_title(soup)
        model, asin_from_page = extract_model_number_and_asin_from_tables(soup)
        if not asin and asin_from_page:
            asin = asin_from_page.strip()

        price_incl = extract_price_tax_included(soup)
        price_excl = 0
        price_incl, price_excl = compute_tax_pair(price_incl, price_excl)

        if not title and not price_incl and not asin:
            if attempt < 3:
                time.sleep(backoff + random.uniform(0.5, 1.4))
                backoff *= 2
                continue
            return None, 'no_data'

        return {
            'supplier': 'Amazon',
            'item_code': asin or '',
            'product_name': title or '',
            'model_number': model or '',
            'price_tax_excluded': str(price_excl) if price_excl else '',
            'price_tax_included': str(price_incl) if price_incl else '',
            'url': url
        }, None

    return None, 'exhausted'

def requests_fetch_pair(url: str, session: requests.Session):
    # HTTP方式のラッパ（将来拡張用）
    return requests_fetch_amazon(url, session)

class AmazonExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Amazon注文情報Excel作成（対策版・exe対応）')
        self.root.geometry('980x780')

        self.mode_var = tk.StringVar(value='new')
        self.use_browser_var = tk.BooleanVar(value=True)     # 既定: ブラウザ方式
        self.headless_var = tk.BooleanVar(value=False)       # 既定: 目視CAPTCHA向け
        self.delay_min_var = tk.DoubleVar(value=3.0)
        self.delay_max_var = tk.DoubleVar(value=6.0)
        self.proxy_var = tk.StringVar(value='')              # 例: http://user:pass@host:port

        self.session = requests.Session()
        self.session.headers.update(amazon_like_headers())

        self.driver = None  # 後片付け用に保持

        self._create_widgets()

    def _create_widgets(self):
        # 1. モード
        mode_frame = ttk.LabelFrame(self.root, text='1. モード選択', padding=10)
        mode_frame.grid(row=0, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        ttk.Radiobutton(mode_frame, text='新規作成', variable=self.mode_var, value='new', command=self._toggle_mode).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(mode_frame, text='既存ファイルに追加', variable=self.mode_var, value='append', command=self._toggle_mode).pack(side=tk.LEFT, padx=10)

        # 2. ファイル設定
        file_frame = ttk.LabelFrame(self.root, text='2. ファイル設定', padding=10)
        file_frame.grid(row=1, column=0, columnspan=3, sticky='ew', padx=10, pady=5)

        self.new_file_frame = ttk.Frame(file_frame)
        self.new_file_frame.pack(side=tk.LEFT, padx=10)

        ttk.Label(self.new_file_frame, text='ファイルパス:').pack(anchor='w')
        nf_row = ttk.Frame(self.new_file_frame)
        nf_row.pack(pady=4)
        self.entry_file_path = ttk.Entry(nf_row, width=50)
        self.entry_file_path.pack(side=tk.LEFT, padx=5)
        ttk.Button(nf_row, text='参照', command=lambda: self._browse(self.entry_file_path, save=True)).pack(side=tk.LEFT)

        ttk.Label(self.new_file_frame, text='シート名:').pack(anchor='w', pady=(8, 0))
        self.entry_sheet_name = ttk.Entry(self.new_file_frame, width=24)
        self.entry_sheet_name.pack()
        self.entry_sheet_name.insert(0, '注文内容')

        self.append_file_frame = ttk.Frame(file_frame)
        ttk.Label(self.append_file_frame, text='既存ファイル:').pack(anchor='w')
        af_row = ttk.Frame(self.append_file_frame)
        af_row.pack(pady=4)
        self.entry_existing_file = ttk.Entry(af_row, width=50, state='disabled')
        self.entry_existing_file.pack(side=tk.LEFT, padx=5)
        self.btn_browse_existing = ttk.Button(af_row, text='参照', command=lambda: self._browse(self.entry_existing_file, save=False), state='disabled')
        self.btn_browse_existing.pack(side=tk.LEFT)

        # 3. 商品入力
        input_frame = ttk.LabelFrame(self.root, text='3. 商品情報入力', padding=10)
        input_frame.grid(row=2, column=0, columnspan=3, sticky='ew', padx=10, pady=5)

        ttk.Label(input_frame, text='Amazon商品URL:').grid(row=0, column=0, sticky='w')
        self.entry_url = ttk.Entry(input_frame, width=70)
        self.entry_url.grid(row=0, column=1, columnspan=2, sticky='ew', padx=5)

        ttk.Label(input_frame, text='個数:').grid(row=1, column=0, sticky='w')
        self.entry_quantity = ttk.Entry(input_frame, width=10)
        self.entry_quantity.grid(row=1, column=1, sticky='w', padx=5)
        self.entry_quantity.insert(0, '1')

        ttk.Button(input_frame, text='リストに追加', command=self._add_to_list).grid(row=1, column=2, padx=5)
        input_frame.columnconfigure(1, weight=1)

        # 4. リスト
        list_frame = ttk.LabelFrame(self.root, text='4. 追加予定の商品リスト', padding=10)
        list_frame.grid(row=3, column=0, columnspan=3, sticky='nsew', padx=10, pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox = tk.Listbox(list_frame, height=10, yscrollcommand=scrollbar.set, font=('Courier', 9))
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)

        btn_frame = ttk.Frame(self.root)
        btn_frame.grid(row=4, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        ttk.Button(btn_frame, text='選択項目を削除', command=self._remove_from_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text='リストをクリア', command=self._clear_list).pack(side=tk.LEFT, padx=5)

        # 5. 対策オプション
        opt_frame = ttk.LabelFrame(self.root, text='5. 対策オプション', padding=10)
        opt_frame.grid(row=5, column=0, columnspan=3, sticky='ew', padx=10, pady=5)
        ttk.Checkbutton(opt_frame, text='ブラウザで取得（推奨）', variable=self.use_browser_var).grid(row=0, column=0, sticky='w')
        ttk.Checkbutton(opt_frame, text='ヘッドレス（CAPTCHA手動解決不可）', variable=self.headless_var).grid(row=0, column=1, sticky='w')
        ttk.Label(opt_frame, text='リクエスト間隔（秒）:').grid(row=1, column=0, sticky='w', pady=(8, 0))
        rng = ttk.Frame(opt_frame)
        rng.grid(row=1, column=1, sticky='w', pady=(8, 0))
        ttk.Entry(rng, width=6, textvariable=self.delay_min_var).pack(side=tk.LEFT)
        ttk.Label(rng, text='〜').pack(side=tk.LEFT, padx=4)
        ttk.Entry(rng, width=6, textvariable=self.delay_max_var).pack(side=tk.LEFT)

        ttk.Label(opt_frame, text='プロキシ（任意）:').grid(row=2, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(opt_frame, width=40, textvariable=self.proxy_var).grid(row=2, column=1, sticky='w', pady=(8, 0))

        # 6. 実行
        run_frame = ttk.Frame(self.root)
        run_frame.grid(row=6, column=0, columnspan=3, sticky='ew', padx=10, pady=10)
        self.btn_run = ttk.Button(run_frame, text='変換実行', command=self._run_conversion)
        self.btn_run.pack(side=tk.LEFT, padx=5)

        # ステータス
        self.status_var = tk.StringVar(value='準備完了')
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=7, column=0, columnspan=3, sticky='ew')

        self.root.rowconfigure(3, weight=1)
        self.root.columnconfigure(0, weight=1)

        self._toggle_mode()

    def _toggle_mode(self):
        if self.mode_var.get() == 'new':
            self.new_file_frame.pack(side=tk.LEFT, padx=10)
            self.append_file_frame.pack_forget()
            self.entry_file_path.config(state='normal')
            self.entry_sheet_name.config(state='normal')
        else:
            self.new_file_frame.pack_forget()
            self.append_file_frame.pack(side=tk.LEFT, padx=10)
            self.entry_existing_file.config(state='normal')
            self.btn_browse_existing.config(state='normal')

    def _browse(self, entry, save=False):
        if save:
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        else:
            file_path = filedialog.askopenfilename(
                filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')]
            )
        if file_path:
            entry.delete(0, tk.END)
            entry.insert(0, file_path)

    def _add_to_list(self):
        url = self.entry_url.get().strip()
        quantity = self.entry_quantity.get().strip()

        if not url:
            messagebox.showwarning('警告', 'URLを入力してください')
            return
        if not is_amazon_product_url(url):
            messagebox.showwarning('警告', 'Amazon.co.jpの商品URLを入力してください\n例) https://www.amazon.co.jp/dp/XXXXXXXXXX')
            return
        if not quantity.isdigit() or int(quantity) < 1:
            messagebox.showwarning('警告', '個数は1以上の数字で入力してください')
            return

        self.listbox.insert(tk.END, f'{url} | 個数: {quantity}')
        self.entry_url.delete(0, tk.END)
        self.entry_quantity.delete(0, tk.END)
        self.entry_quantity.insert(0, '1')
        self.entry_url.focus()

    def _remove_from_list(self):
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning('警告', '削除する項目を選択してください')
            return
        self.listbox.delete(selection)

    def _clear_list(self):
        if messagebox.askyesno('確認', 'リスト内容をすべてクリアしますか？'):
            self.listbox.delete(0, tk.END)

    def _run_conversion(self):
        items = self.listbox.get(0, tk.END)
        if not items:
            messagebox.showwarning('警告', 'URLリストに商品を追加してください')
            return

        self.btn_run.config(state='disabled')
        self.status_var.set('処理中...')
        self.root.update()

        t = threading.Thread(target=self._process_conversion, args=(items,))
        t.daemon = True
        t.start()

    def _process_conversion(self, items):
        use_browser = self.use_browser_var.get()
        headless = self.headless_var.get()
        proxy = self.proxy_var.get().strip() or None

        if proxy:
            self.session.proxies = {'http': proxy, 'https': proxy}

        driver = None
        if use_browser and SEL_AVAILABLE:
            try:
                driver = build_driver(headless=headless, proxy=proxy)
                self.driver = driver
            except Exception as e:
                driver = None
                self.driver = None
                messagebox.showwarning('警告', f'ブラウザ起動に失敗しました。HTTP取得にフォールバックします。\n{e}')

        data_list = []
        total = len(items)
        for idx, item in enumerate(items):
            self.status_var.set(f'処理中... ({idx+1}/{total})')
            self.root.update()

            try:
                parts = item.split(' | 個数: ')
                url = parts[0].strip()
                quantity = int(parts[1].strip())
            except:
                continue

            # ランダムディレイ（人間らしさ）
            dmin = max(0.0, float(self.delay_min_var.get()))
            dmax = max(dmin, float(self.delay_max_var.get()))
            time.sleep(random.uniform(dmin, dmax))

            result = None
            err = None
            if use_browser and driver is not None:
                result, err = selenium_fetch_amazon(driver, url, manual_captcha=(not headless))
                if err in ('bot_block', 'captcha_timeout', 'timeout', 'no_data'):
                    time.sleep(random.uniform(2.0, 4.0))
                    self.session.headers.update(amazon_like_headers())
                    result, err = requests_fetch_pair(url, self.session)
            else:
                self.session.headers.update(amazon_like_headers())
                result, err = requests_fetch_pair(url, self.session)

            if result:
                result['quantity'] = quantity
                data_list.append(result)
            else:
                print(f'取得失敗: {url} ({err})')

        # 後片付け
        try:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    pass
        finally:
            self.driver = None

        if not data_list:
            self.root.after(0, lambda: messagebox.showwarning(
                '警告', '有効な商品情報が取得できませんでした\n（ボット対策やページ構造変更の可能性があります）'
            ))
            self.btn_run.config(state='normal')
            self.status_var.set('準備完了')
            return

        mode = self.mode_var.get()
        try:
            if mode == 'new':
                file_path = self.entry_file_path.get().strip()
                sheet_name = self.entry_sheet_name.get().strip() or '注文内容'
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '新規ファイルパスを入力してください'))
                    return
                ok = self._write_to_excel(file_path, sheet_name, data_list, append=False)
                if ok:
                    self.root.after(0, lambda: messagebox.showinfo('完了', f'Excelファイルに{len(data_list)}件を書き込みました'))
                    self.root.after(0, lambda: self.listbox.delete(0, tk.END))
                else:
                    self.root.after(0, lambda: messagebox.showerror('エラー', 'Excelファイルの書き込みに失敗しました'))
            else:
                file_path = self.entry_existing_file.get().strip()
                if not file_path:
                    self.root.after(0, lambda: messagebox.showwarning('警告', '既存ファイルパスを入力してください'))
                    return
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_name = wb.sheetnames[0] if wb.sheetnames else '注文内容'
                except:
                    sheet_name = '注文内容'
                ok = self._write_to_excel(file_path, sheet_name, data_list, append=True)
                if ok:
                    self.root.after(0, lambda: messagebox.showinfo('完了', f'既存ファイルに{len(data_list)}件を追加しました'))
                    self.root.after(0, lambda: self.listbox.delete(0, tk.END))
                else:
                    self.root.after(0, lambda: messagebox.showerror('エラー', 'Excelファイルの書き込みに失敗しました'))
        finally:
            self.root.after(0, lambda: self.btn_run.config(state='normal'))
            self.root.after(0, lambda: self.status_var.set('準備完了'))

    def _write_to_excel(self, file_path, sheet_name, data_list, append=True):
        try:
            if append and os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                headers = ['仕入元', '商品コード', '商品名', '型番', '単価', '数量', '合計', 'URL', '価格(税込)']
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            for rowd in data_list:
                price_ex = rowd.get('price_tax_excluded', '')
                price_in = rowd.get('price_tax_included', '')
                qty = rowd.get('quantity', '')

                total_ex = ''
                try:
                    if price_ex and qty:
                        total_ex = str(int(price_ex) * int(qty))
                except:
                    total_ex = ''

                row = [
                    rowd.get('supplier', 'Amazon'),
                    rowd.get('item_code', ''),
                    rowd.get('product_name', ''),
                    rowd.get('model_number', ''),
                    price_ex,
                    qty,
                    total_ex,
                    rowd.get('url', ''),
                    price_in
                ]
                ws.append(row)

                # 文字列扱い
                r = ws.max_row
                for col in (2, 4):
                    c = ws.cell(row=r, column=col)
                    c.number_format = '@'
                    c.value = '' if c.value is None else str(c.value)

            # 幅調整
            for column in ws.columns:
                max_len = 0
                letter = get_column_letter(column[0].column)
                for cell in column:
                    v = '' if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[letter].width = min(max_len + 2, 50)

            wb.save(file_path)
            return True
        except Exception as e:
            print('Excel書き込みエラー:', e)
            return False

    # 明示的クリーンアップ（必要なら外部からも呼べる）
    def cleanup_driver(self):
        d = getattr(self, 'driver', None)
        if d is not None:
            try:
                d.quit()
            except Exception:
                pass
            finally:
                self.driver = None

if __name__ == '__main__':
    # exe凍結時の再帰起動（自己実行）を防ぐため、エントリポイントでfreeze_supportを呼ぶ
    import multiprocessing
    multiprocessing.freeze_support()

    try:
        root = tk.Tk()
        app = AmazonExcelApp(root)

        # Ctrl+C で穏当に終了
        root.bind('<Control-c>', lambda e: root.quit())

        def on_close():
            try:
                app.cleanup_driver()
            finally:
                try:
                    root.destroy()
                except:
                    pass

        root.protocol('WM_DELETE_WINDOW', on_close)
        root.mainloop()
    except KeyboardInterrupt:
        try:
            app.cleanup_driver()
        except:
            pass
        try:
            root.destroy()
        except:
            pass
        sys.exit(0)
