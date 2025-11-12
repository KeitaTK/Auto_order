import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import requests
from bs4 import BeautifulSoup
import re
import threading
import time
import json
from urllib.parse import urlparse

class AkizukiExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title('秋月電子 注文情報Excel作成')
        self.root.geometry('900x750')

        self.mode_var = tk.StringVar(value='new')
        self.data_list = []

        # HTTP session with headers
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/120.0.0.0 Safari/537.36'),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ja-JP,ja;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })

        self.create_widgets()

    def create_widgets(self):
        # === モード選択 ===
        mode_frame = ttk.LabelFrame(self.root, text='1. モード選択')
        mode_frame.pack(fill='x', padx=10, pady=8)

        ttk.Radiobutton(mode_frame, text='新規Excel作成', variable=self.mode_var,
                        value='new', command=self.update_mode_state).pack(side='left', padx=5, pady=5)
        ttk.Radiobutton(mode_frame, text='既存Excelに追記', variable=self.mode_var,
                        value='append', command=self.update_mode_state).pack(side='left', padx=5, pady=5)

        # === ファイル設定 ===
        file_frame = ttk.LabelFrame(self.root, text='2. ファイル設定')
        file_frame.pack(fill='x', padx=10, pady=8)

        # 新規作成
        new_frame = ttk.Frame(file_frame)
        new_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(new_frame, text='保存先:').grid(row=0, column=0, sticky='w')
        self.save_path_var = tk.StringVar()
        self.save_path_entry = ttk.Entry(new_frame, textvariable=self.save_path_var, width=70)
        self.save_path_entry.grid(row=0, column=1, padx=5)
        self.save_path_btn = ttk.Button(new_frame, text='参照', command=self.browse_save_path)
        self.save_path_btn.grid(row=0, column=2, padx=5)

        ttk.Label(new_frame, text='シート名:').grid(row=1, column=0, sticky='w', pady=(6,0))
        self.sheet_name_var = tk.StringVar(value='注文内容')
        self.sheet_name_entry = ttk.Entry(new_frame, textvariable=self.sheet_name_var, width=30)
        self.sheet_name_entry.grid(row=1, column=1, sticky='w', pady=(6,0))

        # 追記
        append_frame = ttk.Frame(file_frame)
        append_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(append_frame, text='既存Excel:').grid(row=0, column=0, sticky='w')
        self.append_file_var = tk.StringVar()
        self.append_file_entry = ttk.Entry(append_frame, textvariable=self.append_file_var, width=70)
        self.append_file_entry.grid(row=0, column=1, padx=5)
        self.append_file_btn = ttk.Button(append_frame, text='参照', command=self.browse_existing_file)
        self.append_file_btn.grid(row=0, column=2, padx=5)

        # === 商品入力 ===
        input_frame = ttk.LabelFrame(self.root, text='3. 商品情報入力（秋月電子の商品ページURLと数量）')
        input_frame.pack(fill='x', padx=10, pady=8)
        ttk.Label(input_frame, text='商品URL:').grid(row=0, column=0, sticky='w')
        self.url_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.url_var, width=75).grid(row=0, column=1, padx=5, pady=5, sticky='w')

        ttk.Label(input_frame, text='数量:').grid(row=1, column=0, sticky='w')
        self.qty_var = tk.StringVar(value='1')
        ttk.Entry(input_frame, textvariable=self.qty_var, width=10).grid(row=1, column=1, sticky='w', pady=5)
        ttk.Button(input_frame, text='リストに追加', command=self.add_to_list).grid(row=1, column=2, padx=10, pady=5, sticky='w')

        # === リスト ===
        list_frame = ttk.LabelFrame(self.root, text='4. 追加予定リスト')
        list_frame.pack(fill='both', expand=True, padx=10, pady=8)
        self.listbox = tk.Listbox(list_frame, height=15)
        self.listbox.pack(side='left', fill='both', expand=True, padx=(5,0), pady=5)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.listbox.yview)
        scrollbar.pack(side='left', fill='y', pady=5)
        self.listbox.config(yscrollcommand=scrollbar.set)

        btns = ttk.Frame(list_frame)
        btns.pack(side='left', fill='y', padx=8)
        ttk.Button(btns, text='選択削除', command=self.remove_selected).pack(fill='x', pady=4)
        ttk.Button(btns, text='全クリア', command=self.clear_list).pack(fill='x', pady=4)

        # === 実行 ===
        run_frame = ttk.LabelFrame(self.root, text='5. 実行')
        run_frame.pack(fill='x', padx=10, pady=8)
        ttk.Button(run_frame, text='変換実行', command=self.run_conversion).pack(side='left', padx=5, pady=5)

        # === ステータス ===
        self.status_var = tk.StringVar(value='準備完了')
        status = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', anchor='w')
        status.pack(fill='x', padx=10, pady=(0,8))

        self.update_mode_state()

    def update_mode_state(self):
        mode = self.mode_var.get()
        new_enable = (mode == 'new')
        append_enable = (mode == 'append')

        for w, enable in [
            (self.save_path_entry, new_enable),
            (self.save_path_btn, new_enable),
            (self.sheet_name_entry, True),  # 追記時もシート名指定できるよう常時有効
            (self.append_file_entry, append_enable),
            (self.append_file_btn, append_enable),
        ]:
            try:
                w.config(state=('normal' if enable else 'disabled'))
            except Exception:
                pass

    def browse_save_path(self):
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel Workbook', '*.xlsx')]
        )
        if path:
            self.save_path_var.set(path)

    def browse_existing_file(self):
        path = filedialog.askopenfilename(
            filetypes=[('Excel Workbook', '*.xlsx')]
        )
        if path:
            self.append_file_var.set(path)

    def add_to_list(self):
        url = self.url_var.get().strip()
        qty_str = self.qty_var.get().strip()

        if not url:
            messagebox.showwarning('警告', '商品URLを入力してください。')
            return
        if not self.is_akizuki_url(url):
            messagebox.showwarning('警告', '秋月電子通商のURL（akizukidenshi.com）を入力してください。')
            return
        if not qty_str.isdigit() or int(qty_str) < 1:
            messagebox.showwarning('警告', '数量は1以上の整数で入力してください。')
            return

        item_text = f'{url} | 個数: {qty_str}'
        self.listbox.insert('end', item_text)
        self.url_var.set('')
        self.qty_var.set('1')

    def remove_selected(self):
        selected = list(self.listbox.curselection())
        for i in reversed(selected):
            self.listbox.delete(i)

    def clear_list(self):
        self.listbox.delete(0, 'end')

    def set_status(self, text):
        self.status_var.set(text)
        self.root.update_idletasks()

    def run_conversion(self):
        count = self.listbox.size()
        if count == 0:
            messagebox.showwarning('警告', 'リストに商品がありません。')
            return

        mode = self.mode_var.get()
        if mode == 'new':
            if not self.save_path_var.get().strip():
                messagebox.showwarning('警告', '保存先のExcelファイルパスを指定してください。')
                return
            if not self.sheet_name_var.get().strip():
                messagebox.showwarning('警告', 'シート名を入力してください。')
                return
        else:
            if not self.append_file_var.get().strip():
                messagebox.showwarning('警告', '追記先のExcelファイルを指定してください。')
                return

        self.set_status('処理を開始します...')
        thread = threading.Thread(target=self.worker_process, daemon=True)
        thread.start()

    def worker_process(self):
        try:
            items = []
            for i in range(self.listbox.size()):
                text = self.listbox.get(i)
                m = re.match(r'(.+?)\s+\|\s+個数:\s*(\d+)', text)
                if not m:
                    continue
                url = m.group(1).strip()
                qty = int(m.group(2))
                self.set_status(f'取得中: {url}')

                product = self.fetch_and_parse(url)
                if product is None:
                    continue
                product['quantity'] = qty
                items.append(product)
                time.sleep(0.7)  # polite interval

            if not items:
                messagebox.showwarning('警告', '情報を取得できませんでした。ページ構造の変更やメンテナンスの可能性があります。')
                self.set_status('準備完了')
                return

            mode = self.mode_var.get()
            if mode == 'new':
                self.write_new_excel(self.save_path_var.get().strip(),
                                     self.sheet_name_var.get().strip(),
                                     items)
                messagebox.showinfo('完了', f'Excelを作成しました:\n{self.save_path_var.get().strip()}')
            else:
                sheet = self.sheet_name_var.get().strip() or '注文内容'
                self.append_to_excel(self.append_file_var.get().strip(),
                                     sheet,
                                     items)
                messagebox.showinfo('完了', f'Excelに追記しました:\n{self.append_file_var.get().strip()}')

            self.clear_list()
            self.set_status('準備完了')

        except Exception as e:
            messagebox.showerror('エラー', f'処理中にエラーが発生しました:\n{e}')
            self.set_status('準備完了')

    # ========== Scraping ==========

    def is_akizuki_url(self, url: str) -> bool:
        try:
            p = urlparse(url)
            host = (p.netloc or '').lower()
            return 'akizukidenshi.com' in host
        except Exception:
            return False

    def fetch_and_parse(self, url: str):
        html = self.fetch_page(url)
        if not html:
            return None
        soup = BeautifulSoup(html, 'html.parser')

        # Extract fields
        name = self.extract_name(soup)
        model = self.extract_model(soup)
        item_code = self.extract_item_code(soup, url)
        price_ex, price_in = self.extract_prices(soup)

        # Fallback: JSON-LD price/name if missing
        if (price_ex is None or price_in is None) or not name:
            jl_name, jl_price = self.extract_jsonld(soup)
            if not name and jl_name:
                name = jl_name
            if price_in is None and jl_price is not None:
                price_in = jl_price

        # Normalize to int
        price_ex = self.to_int(price_ex)
        price_in = self.to_int(price_in)

        # Derive missing price using 10% tax
        if price_in is None and price_ex is not None:
            price_in = int(round(price_ex * 1.1))
        if price_ex is None and price_in is not None:
            price_ex = int(round(price_in / 1.1))

        if not any([name, model, item_code, price_ex, price_in]):
            return None

        return {
            'supplier': '秋月電子通商',
            'item_code': item_code or '',
            'name': name or '',
            'model': model or '',
            'price_excl_tax': price_ex if price_ex is not None else 0,
            'price_incl_tax': price_in if price_in is not None else 0,
            'url': url,
        }

    def fetch_page(self, url: str, retries: int = 3, timeout: int = 20):
        last_err = None
        for i in range(retries):
            try:
                r = self.session.get(url, timeout=timeout)
                if r.status_code == 200 and 'text/html' in r.headers.get('Content-Type', '').lower():
                    text = r.text
                    if any(x in text for x in ['アクセスが集中', 'メンテナンス', 'ただいま処理中']):
                        time.sleep(1.2 + i * 0.5)
                        continue
                    return text
                else:
                    last_err = f'HTTP {r.status_code}'
            except Exception as e:
                last_err = str(e)
            time.sleep(1.0 + i * 0.5)
        return None

    def extract_name(self, soup: BeautifulSoup):
        # Priority: h1 product name
        el = soup.select_one('h1.h1-goods-name')  # div.block-goods-name > h1
        if el and el.get_text(strip=True):
            return el.get_text(strip=True)

        og = soup.find('meta', property='og:title')
        if og and og.get('content'):
            return og['content'].strip()

        if soup.title and soup.title.string:
            title = soup.title.string.strip()
            title = re.sub(r'\s*[｜|:].*$', '', title)
            return title
        return None

    def extract_model(self, soup: BeautifulSoup):
        # Direct id
        dd = soup.select_one('dd#spec_number')
        if dd and dd.get_text(strip=True):
            return dd.get_text(strip=True)

        # dt "型番" -> sibling dd
        for dt in soup.select('dt, th'):
            label = dt.get_text(strip=True)
            if '型番' in label or '型式' in label or '品番' in label:
                sib = None
                if dt.name == 'dt':
                    sib = dt.find_next_sibling('dd')
                else:
                    sib = dt.find_next_sibling('td')
                if sib and sib.get_text(strip=True):
                    return sib.get_text(strip=True)

        # Inline fallback
        text = soup.get_text('\n', strip=True)
        m = re.search(r'(?:型番|型式|品番)\s*[:：]\s*([^\s　|｜\n]+)', text)
        if m:
            return m.group(1)
        return None

    def extract_item_code(self, soup: BeautifulSoup, url: str):
        # DOM first
        dd = soup.select_one('dd#spec_goods')
        if dd and dd.get_text(strip=True):
            return dd.get_text(strip=True)

        for dt in soup.select('dt, th'):
            if '販売コード' in dt.get_text(strip=True) or '商品コード' in dt.get_text(strip=True):
                sib = None
                if dt.name == 'dt':
                    sib = dt.find_next_sibling('dd')
                else:
                    sib = dt.find_next_sibling('td')
                if sib and sib.get_text(strip=True):
                    return sib.get_text(strip=True)

        # URL pattern: /catalog/g/g123456/
        try:
            path = urlparse(url).path
            m = re.search(r'/catalog/g/g(\d+)/?', path)
            if m:
                return m.group(1)
        except Exception:
            pass

        return None

    def extract_prices(self, soup: BeautifulSoup):
        price_incl = None
        price_excl = None

        # Explicit elements
        incl_el = soup.select_one('.block-goods-price--price')
        if incl_el:
            m = re.search(r'￥\s*([0-9,]+)', incl_el.get_text(' ', strip=True))
            if m:
                price_incl = self.to_int(m.group(1))

        excl_el = soup.select_one('.block-goods-price--net-price')
        if excl_el:
            m = re.search(r'￥\s*([0-9,]+)', excl_el.get_text(' ', strip=True))
            if m:
                price_excl = self.to_int(m.group(1))

        # Context scan with 税込/税抜 hints
        if price_incl is None or price_excl is None:
            text = soup.get_text('\n', strip=True)
            for m in re.finditer(r'￥\s*([0-9,]+)', text):
                val = self.to_int(m.group(1))
                if val is None:
                    continue
                s = max(0, m.start() - 15)
                e = min(len(text), m.end() + 15)
                win = text[s:e]
                if price_incl is None and ('税込' in win or '(税込' in win or '（税込' in win):
                    price_incl = val
                if price_excl is None and ('税抜' in win or '税別' in win):
                    price_excl = val

        return price_excl, price_incl

    def extract_jsonld(self, soup: BeautifulSoup):
        # Returns (name, price_incl_tax)
        try:
            for sc in soup.find_all('script', {'type': 'application/ld+json'}):
                content = sc.string or sc.get_text()
                if not content:
                    continue
                data = json.loads(content)
                # Handle list or single object
                if isinstance(data, list):
                    candidates = data
                else:
                    candidates = [data]
                for obj in candidates:
                    if isinstance(obj, dict) and obj.get('@type') in ('Product', 'product', 'Offer'):
                        name = obj.get('name')
                        price = None
                        if 'offers' in obj and isinstance(obj['offers'], dict):
                            price = obj['offers'].get('price')
                        elif 'price' in obj:
                            price = obj.get('price')
                        return (name, self.to_int(price))
        except Exception:
            pass
        return (None, None)

    def to_int(self, v):
        if v is None:
            return None
        if isinstance(v, int):
            return v
        try:
            s = str(v)
            s = re.sub(r'[^\d]', '', s)
            if not s:
                return None
            return int(s)
        except Exception:
            return None

    # ========== Excel Writing ==========

    def write_new_excel(self, path: str, sheet_name: str, items: list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = ['仕入元', '商品コード', '商品名', '型番', '単価(税別)', '数量', '合計(税別)', 'URL', '価格(税込)']
        ws.append(headers)

        # Style header
        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='DDDDDD')
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Rows
        for it in items:
            price_ex = it.get('price_excl_tax', 0) or 0
            qty = it.get('quantity', 0) or 0
            total_ex = price_ex * qty
            row = [
                it.get('supplier', ''),
                str(it.get('item_code', '')),
                it.get('name', ''),
                str(it.get('model', '')),
                price_ex,
                qty,
                total_ex,
                it.get('url', ''),
                it.get('price_incl_tax', 0) or 0,
            ]
            ws.append(row)

        # Column widths
        self.autofit_columns(ws, max_width=50)

        wb.save(path)

    def append_to_excel(self, path: str, sheet_name: str, items: list):
        try:
            wb = openpyxl.load_workbook(path)
        except FileNotFoundError:
            self.write_new_excel(path, sheet_name, items)
            return

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            headers = ['仕入元', '商品コード', '商品名', '型番', '単価(税別)', '数量', '合計(税別)', 'URL', '価格(税込)']
            ws.append(headers)
            header_font = Font(bold=True)
            header_fill = PatternFill('solid', fgColor='DDDDDD')
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for it in items:
            price_ex = it.get('price_excl_tax', 0) or 0
            qty = it.get('quantity', 0) or 0
            total_ex = price_ex * qty
            row = [
                it.get('supplier', ''),
                str(it.get('item_code', '')),
                it.get('name', ''),
                str(it.get('model', '')),
                price_ex,
                qty,
                total_ex,
                it.get('url', ''),
                it.get('price_incl_tax', 0) or 0,
            ]
            ws.append(row)

        self.autofit_columns(ws, max_width=50)
        wb.save(path)

    def autofit_columns(self, ws, max_width=50):
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for idx, val in enumerate(row, start=1):
                length = len(str(val)) if val is not None else 0
                widths[idx] = max(widths.get(idx, 0), length)
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 10), max_width)


def main():
    root = tk.Tk()
    app = AkizukiExcelApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()
